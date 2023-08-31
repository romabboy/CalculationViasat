from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import os, time


def calculat(packet, date):
    date = int(date)
    SOCIALNIY = 49
    NATIONALNIY = 99
    SIMEYNIY = 99
    PRESTIGEHDPREMIUM = 249
    TWIN1 = 69
    PRESTIGE = 199
    #
    # if date > 7:
    #     SIMEYNIY = 99
    #     PRESTIGEHDPREMIUM = 125
    #     PRESTIGE = 99
    # if date > 8:
    #     NATIONALNIY = 99
    #     SOCIALNIY = 99



    obj = {
        "SOCIALNIY": (SOCIALNIY * 1.5)*0.9,
        "NATIONALNIY": (NATIONALNIY * 1.5)*0.9,
        "SIMEYNIY": (SIMEYNIY * 1.5)*0.9,
        "PRESTIGE HD PREMIUM": (PRESTIGEHDPREMIUM * 1.5)*0.9,
        "TWIN1": (TWIN1 * 1.5)*0.9,
        "TWIN2": 0,
        "TWIN3": 0,
        "PRESTIGE": (PRESTIGE * 1.5)*0.9,
    }

    sum = obj.get(packet,None)
    return sum

def monthStr(num):
    obj = {
        "1":"Січень",
        "2": "Лютий",
        "3": "Березень",
        "4": "Квітень",
        "5": "Травень",
        "6": "Червень",
        "7": "Липень",
        "8": "Серпень",
        "9": "Вересень",
        "10": "Жовтень",
        "11": "Листопад",
        "12": "Грудень",
    }


    return obj[str(num)]

def alignmentHorizontal(sheet,row=1,column=1):
    A1 = sheet.cell(row,column)
    A1.alignment = Alignment(horizontal='center')


def addInfo(sheet,month):
    sheet.append([monthStr(month)])
    sheet.merge_cells("A1:K1")
    alignmentHorizontal(sheet)
    Linfo = ["Дата","Місяць","Особовий рахунок","Номер майстра","ПІБ","Пакет","Компанія підключення",
             "Номер тюнера","Перша оплата","Статус","Сума"]
    sheet.append(Linfo)

    for i in range(len(Linfo)):
        alignmentHorizontal(sheet,row=2,column=i+1)

def autoScale(book):
    ws = list(book.worksheets)

    for sheet in ws:
        word = "A"
        for i,cel in enumerate(sheet.columns,1):
            size = sheet.cell(column=i,row=3).value
            size = len(str(size)) + 3
            sheet.column_dimensions[word].width = size
            word = chr(ord(word)+1)

def createWB(arr):
    objFromArr = {}
    Masters_Name = ""
    LBook = []
    date = time.localtime(time.time())
    date = list(date)
    date = f"{date[2]} {date[1]} {date[0]}"
    pathMkDir = fr"C:\Users\Dell\Desktop\Робота\PythonCalculate\{date}"
    os.mkdir(pathMkDir)

    for row in arr:
        if Masters_Name != row[4]:
            Masters_Name = row[4]
            objFromArr[Masters_Name] = {}

        month = row[1]
        if not objFromArr[Masters_Name].get(month):
            objFromArr[Masters_Name][month] = {'data': [row], 'sum': 0}
            objFromArr[Masters_Name][month]["sum"] = row[-1]
        else:
            objFromArr[Masters_Name][month]["data"].append(row)
            objFromArr[Masters_Name][month]["sum"] += row[-1]

    for name in objFromArr:
        book = Workbook()
        book.name = name
        book.remove(book.active)
        monthSort = list(objFromArr[name].keys())
        monthSort.sort()

        for month in monthSort:

            sheet = book.create_sheet(monthStr(month))
            addInfo(sheet,month)

            for masterData in objFromArr[name][month]["data"]:
                sheet.append(masterData)

            # style
            SUM = sheet.cell(column = sheet.max_column, row = sheet.max_row+1)
            SUM.value = objFromArr[name][month]["sum"]
            SUM.font = Font(bold="b")

        LBook.append(book)

        for book in LBook:
            autoScale(book)



    for file in LBook:
        file.save(fr"{pathMkDir}\{file.name}.xlsx")
